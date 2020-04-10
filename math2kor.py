# How to run
# python math2kor.py
# or
# python math2kor.py YourExelFile.xlsx

from openpyxl import load_workbook, Workbook
from TexSoup import TexSoup
import pyjosa
import sys
import re

class Eq2Script:
    def __init__(self):
        self.math_table = self.xlsx2dict('math_table.xlsx')
        self.math_table2 = self.xlsx2dict2('math_table2.xlsx')

    # 엑셀을 딕셔너리로 변환
    def xlsx2dict(self, xlsx):
        load_wb = load_workbook(xlsx, data_only=True)
        load_ws = load_wb['Sheet1']

        math_table = {}

        for row in load_ws.rows:
            
            # 수식
            equation = row[0].value
            # 스크립트
            script = row[1].value

            math_table[equation] = script
        return math_table

    def xlsx2dict2(self, xlsx):
        load_wb = load_workbook(xlsx, data_only=True)
        load_ws = load_wb['Sheet1']

        math_table = {}

        for row in load_ws.rows:
            # 수식
            node_name = row[0].value
            # 스크립트
            script1 = row[1].value
            script2 = row[2].value

            math_table[node_name] = (script1, script2)
        return math_table

    def convert_wrong_char(self, raw_text):
        return raw_text.replace('￦','\\')

    def post_process(self, text):
        text = text.replace('\\','').replace('.','').replace('(','').replace(')','')
        text = text.replace('를','(을)를').replace('와','(와)과').replace('는','(은)는')
        while text.count('  '):
            text = text.replace('  ',' ')
        return text

    def textree(self, node): 
        script = ''
        prev = None
        if node.name == 'frac': 
            script += self.textree(TexSoup(node.args[1].value)) # node.args는 텍스트, 이를 Node화 시키기 위해 TexSoup 사용
            script += '분의'
            script += self.textree(TexSoup(node.args[0].value))
        elif node.name == 'sqrt':
            if len(node.args) == 1 :
                script += '루트'
                script += self.textree(TexSoup(node.args[0].value))
            elif node.args[0].value == str(2) :
                script += '루트'
                script += self.textree(TexSoup(node.args[1].value))
            elif node.args[0].value == str(3) :
                script += self.textree(TexSoup(node.args[1].value)) 
                script += '의 세제곱근'
            elif node.args[0].value == str(4) :
                script += self.textree(TexSoup(node.args[1].value)) 
                script += '의 네제곱근'
            else:
                script += self.textree(TexSoup(node.args[1].value)) 
                script += '의 '
                script += self.textree(TexSoup(node.args[0].value))
                script += '제곱근'
        elif node.name in self.math_table2:
            script += self.textree(TexSoup(node.args[1].value)) # node.args는 텍스트, 이를 Node화 시키기 위해 TexSoup 사용
            script += self.math_table2[node.name][0]
            script += self.textree(TexSoup(node.args[0].value))
            script += self.math_table2[node.name][1]
        elif node.name == '.':
            script += self.textree(TexSoup(node.args[0].value))
            script += '땡'
            
        elif node.name != '[tex]':
            # \pm, \sqrt 등을 변환
            try:
                script += self.math_table.get(str(node.name))
            except:
                # 변환 테이블에 없을 경우 기호 이름 반환
                print(node.name)
            # \sqrt{...}에서 괄호 안에 있는 것을 변환
            for arg in node.args:
                script += self.textree(TexSoup(arg.value))
        else:
            for cont in node.contents: # node || token 으로만 나누어짐 // '\'가 있으면 node, 없으면 token
                if type(cont) == type(node):
                    script += self.textree(cont)
                # 텍스트면 그냥 출력
                # elif str(type(cont)) == "<class 'TexSoup.data.RArg'>":
                elif prev is not None:
                    if str(cont) == "{2}":
                        script += "제곱"
                    elif str(cont)== "{3}":
                        script += "세제곱"
                    else:    
                        script += '의 '
                        for c_raw in cont:
                            c = str(c_raw)
                            script += self.textree(TexSoup(c))
                        script += ' 승'        
                    prev = None
                else:
                    for c_raw in cont:
                        c = str(c_raw)
                        # 특수 문자 인지 확인
                        if c == "'":
                            script += '프라임'
                        elif c in '^_':
                            prev = c                      
                        # 문자가 table에 있으면 변환
                        elif c in self.math_table:
                            script += self.math_table.get(c)
                        # table에 없으면 그대로 출력
                        else:
                            script += c
        return script

    def script(self, equation):
        if equation.find('>') >= 0 or equation.find('<') >= 0:
            equation=equation.replace('>','\>').replace('<','\<')
        
        for key in (self.math_table2.keys()) :
            if key in equation :
                tmp = equation.split('\\'+key)
                equation = '\\' + key +'{'+tmp[1]+'}{'+tmp[0]+'}'
        node = TexSoup(equation)
        script = self.textree(node)
        return script

    def text2script(self, raw_text):
        text = ''
        for i, t in enumerate(self.convert_wrong_char(raw_text).split('$')):
            if (i % 2) == 0:
                text += t
            else:
                text += self.script(t)
        text = self.post_process(text)
        return text

    def xlsx2script(self, xlsx):
        load_wb = load_workbook(xlsx, data_only=True)
        load_ws = load_wb.active

        if load_ws.title + '_notex' in load_wb.sheetnames:
            save_ws = load_wb[load_ws.title + '_notex']
        else:
            save_ws = load_wb.create_sheet(load_ws.title + '_notex')
        for row_id, row in enumerate(load_ws.rows):
            for collumn_id, collumn in enumerate(row):
                raw_text = collumn.value
                if raw_text is not None: 
                    save_ws.cell(row_id+1, collumn_id+1, self.text2script(raw_text))

        load_wb.save(xlsx)
        print('저장완료...')

                
if __name__ == '__main__':
    
    tcs = open('test_cases.txt','rt',encoding='utf8') 

    for idx,tc in enumerate(tcs):
        sample = Eq2Script().text2script(tc)
        sample = pyjosa.replace_josa(sample)
        print('({}){}'.format(idx+1,sample))

